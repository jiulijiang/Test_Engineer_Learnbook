from time import sleep
from selenium import webdriver
from selenium.common import NoAlertPresentException, StaleElementReferenceException
from selenium.webdriver.common.by import By
from can_reuesu_code.login import login_byhy
import openpyxl


def del_element(wd: webdriver, element_css_selector: str):
    """
    自动点击传入的所有删除按钮
    :param wd:
    :param element_css_selector:
    :return:
    """
    while True:
        try:
            # 每次循环都重新查找元素
            element_delete_btns = wd.find_elements(By.CSS_SELECTOR, element_css_selector)
            if not element_delete_btns:
                break  # 没有删除按钮则退出循环

            # 点击首项删除按钮
            element_delete_btns[0].click()

            # 处理确认删除弹窗
            try:
                wd.switch_to.alert.accept()  # 确定删除
            except NoAlertPresentException:
                pass  # 没有弹窗则继续

            # 等待一段时间让页面更新
            sleep(0.5)

        except StaleElementReferenceException:
            # 如果元素过期，等待后重新尝试
            sleep(0.5)
            continue
        except Exception as e:
            # 处理其他可能的异常
            print(f"删除元素时出现异常: {e}")
            break





wd = webdriver.Edge()
login_byhy(wd, "http://127.0.0.1:8080")
wd.implicitly_wait(2)

# 初始化信息
wd.find_element(By.CSS_SELECTOR, 'a[href="#/orders"]').click() #切换到订单界面
del_element(wd,'.btn-group-sm label')
wd.find_element(By.CSS_SELECTOR, 'a[href="#/medicines"]').click() #切换到药品界面
del_element(wd,'.btn-xs:nth-child(2)')
wd.find_element(By.CSS_SELECTOR, 'a[href="#/customers"]').click() #切换到客户界面
del_element(wd,'.btn-xs:nth-child(2)')

#存储新信息
# 切换到药品管理页面
wd.find_element(By.CSS_SELECTOR, 'a[href="#/medicines"]').click()

# 使用 openpyxl 读取药品数据
medicines_path = r"C:\Users\thoma\Desktop\downloads\medicines.xlsx"
wb = openpyxl.load_workbook(medicines_path, read_only=True)
sheet = wb.active  # 获取第一个工作表

wd.find_element(By.CSS_SELECTOR, 'button[class="btn btn-green btn-outlined btn-md"]').click()  # 触发新增按钮

# 遍历 Excel 行 (从第一行开始)
for row in sheet.iter_rows(min_row=1, values_only=True):
    wd.find_element(By.CSS_SELECTOR, 'div[class="col-lg-8 col-md-8 col-sm-8"] > div:nth-child(1)>input').send_keys(row[0])
    wd.find_element(By.CSS_SELECTOR, 'div[class="col-lg-8 col-md-8 col-sm-8"] > div:nth-child(2)>input').send_keys(row[1])
    wd.find_element(By.CSS_SELECTOR, 'div[class="col-lg-8 col-md-8 col-sm-8"] > div:nth-child(3)>textarea').send_keys(row[2])
    wd.find_element(By.XPATH, '//button[@class="btn btn-green btn-outlined btn-xs"][1]').click()  # 添加
    sleep(1)
wb.close()  # 关闭工作簿

# 切换到客户界面
wd.find_element(By.CSS_SELECTOR, 'a[href="#/customers"]').click()

# 使用 openpyxl 读取客户数据
customers_path = r"C:\Users\thoma\Desktop\downloads\customers.xlsx"
wb = openpyxl.load_workbook(customers_path, read_only=True)
sheet = wb.active  # 获取第一个工作表

wd.find_element(By.CSS_SELECTOR, 'button[class="btn btn-green btn-outlined btn-md"]').click()  # 触发新增按钮

# 遍历 Excel 行 (从第一行开始)
for row in sheet.iter_rows(min_row=1, values_only=True):
    wd.find_element(By.CSS_SELECTOR, 'div[class="col-lg-8 col-md-8 col-sm-8"] > div:nth-child(1)>input').send_keys(row[0])
    wd.find_element(By.CSS_SELECTOR, 'div[class="col-lg-8 col-md-8 col-sm-8"] > div:nth-child(2)>input').send_keys(row[1])
    wd.find_element(By.CSS_SELECTOR, 'div[class="col-lg-8 col-md-8 col-sm-8"] > div:nth-child(3)>textarea').send_keys(row[2])
    wd.find_element(By.XPATH, '//button[@class="btn btn-green btn-outlined btn-xs"][1]').click()  # 添加
    sleep(1)
wb.close()  # 关闭工作簿

# 切换到订单页面并创建订单
wd.find_element(By.CSS_SELECTOR, 'a[href="#/orders"]').click()
wd.find_element(By.CSS_SELECTOR, 'button[class="btn btn-green btn-outlined btn-md"]').click()
wd.find_element(By.CSS_SELECTOR, 'div.col-lg-8>div:nth-child(1) input').send_keys('自动化测试订单')
wd.find_element(By.CSS_SELECTOR, 'div.col-lg-8>div:nth-child(2)>input').send_keys('南京中医院2\n')
wd.find_element(By.CSS_SELECTOR, 'div.col-lg-8>div:nth-child(2)>select>option').click()
wd.find_element(By.CSS_SELECTOR, 'div.col-lg-8>div:nth-child(3)>input').send_keys('青霉素盒装1')
wd.find_element(By.CSS_SELECTOR, 'div.col-lg-8>div:nth-child(3)>select>option:nth-child(1)').click()
wd.find_element(By.CSS_SELECTOR, "input[type='number'").send_keys('100')
wd.find_element(By.CSS_SELECTOR, '.col-lg-12>div:nth-child(3)>button:nth-child(1)').click()  # 添加