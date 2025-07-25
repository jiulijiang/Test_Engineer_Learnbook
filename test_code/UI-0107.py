from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from can_reuesu_code.login import login_byhy
import openpyxl  # 替换 xlrd 为 openpyxl

wd = webdriver.Edge()
login_byhy(wd, "http://127.0.0.1:8080")
wd.implicitly_wait(5)

# 切换到药品管理页面
wd.find_element(By.CSS_SELECTOR, 'a[href="#/medicines"]').click()

# 使用 openpyxl 读取药品数据
medicines_path = r"C:\Users\thoma\Desktop\downloads\medicines.xlsx"
wb = openpyxl.load_workbook(medicines_path, read_only=True)
sheet = wb.active  # 获取第一个工作表

wd.find_element(By.CSS_SELECTOR, 'button[class="btn btn-green btn-outlined btn-md"]').click()  # 触发新增按钮

# 遍历 Excel 行 (从第一行开始)
for row in sheet.iter_rows(min_row=1, values_only=True):
    wd.find_element(By.CSS_SELECTOR, 'div[class="col-lg-8 col-md-8 col-sm-8"] > div:nth-child(1)>input').send_keys(row[0])
    wd.find_element(By.CSS_SELECTOR, 'div[class="col-lg-8 col-md-8 col-sm-8"] > div:nth-child(2)>input').send_keys(row[1])
    wd.find_element(By.CSS_SELECTOR, 'div[class="col-lg-8 col-md-8 col-sm-8"] > div:nth-child(3)>textarea').send_keys(row[2])
    wd.find_element(By.XPATH, '//button[@class="btn btn-green btn-outlined btn-xs"][1]').click()  # 添加
    sleep(1)
wb.close()  # 关闭工作簿

# 切换到客户界面
wd.find_element(By.CSS_SELECTOR, 'a[href="#/customers"]').click()

# 使用 openpyxl 读取客户数据
customers_path = r"C:\Users\thoma\Desktop\downloads\customers.xlsx"
wb = openpyxl.load_workbook(customers_path, read_only=True)
sheet = wb.active  # 获取第一个工作表

wd.find_element(By.CSS_SELECTOR, 'button[class="btn btn-green btn-outlined btn-md"]').click()  # 触发新增按钮

# 遍历 Excel 行 (从第一行开始)
for row in sheet.iter_rows(min_row=1, values_only=True):
    wd.find_element(By.CSS_SELECTOR, 'div[class="col-lg-8 col-md-8 col-sm-8"] > div:nth-child(1)>input').send_keys(row[0])
    wd.find_element(By.CSS_SELECTOR, 'div[class="col-lg-8 col-md-8 col-sm-8"] > div:nth-child(2)>input').send_keys(row[1])
    wd.find_element(By.CSS_SELECTOR, 'div[class="col-lg-8 col-md-8 col-sm-8"] > div:nth-child(3)>textarea').send_keys(row[2])
    wd.find_element(By.XPATH, '//button[@class="btn btn-green btn-outlined btn-xs"][1]').click()  # 添加
    sleep(1)
wb.close()  # 关闭工作簿

# 切换到订单页面并创建订单
wd.find_element(By.CSS_SELECTOR, 'a[href="#/orders"]').click()
wd.find_element(By.CSS_SELECTOR, 'button[class="btn btn-green btn-outlined btn-md"]').click()
wd.find_element(By.CSS_SELECTOR, 'div.col-lg-8>div:nth-child(1) input').send_keys('自动化测试订单')
wd.find_element(By.CSS_SELECTOR, 'div.col-lg-8>div:nth-child(2)>input').send_keys('南京中医院2\n')
wd.find_element(By.CSS_SELECTOR, 'div.col-lg-8>div:nth-child(2)>select>option').click()
wd.find_element(By.CSS_SELECTOR, 'div.col-lg-8>div:nth-child(3)>input').send_keys('青霉素盒装1')
wd.find_element(By.CSS_SELECTOR, 'div.col-lg-8>div:nth-child(3)>select>option:nth-child(1)').click()
wd.find_element(By.CSS_SELECTOR, "input[type='number'").send_keys('100')
wd.find_element(By.CSS_SELECTOR, '.col-lg-12>div:nth-child(3)>button:nth-child(1)').click()  # 添加